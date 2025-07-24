import pandas as pd
from openpyxl import load_workbook

# Rutas de los archivos
archivo_1 = r'C:\Users\Mateo Zapata\OneDrive\Escritorio\Automatizaciones\2. Pedidos y Órdenes de Carga\Descargas Maestros Endpoints\listado_terceros.xlsx'  # Archivo principal a complementar
archivo_2 = r'C:\Users\Mateo Zapata\OneDrive\Escritorio\Automatizaciones\2. Pedidos y Órdenes de Carga\Descargas Maestros Endpoints\listado_terceros (1).xlsx'    # Archivo con nueva información
archivo_3 = r'C:\Users\Mateo Zapata\OneDrive\Escritorio\Automatizaciones\2. Pedidos y Órdenes de Carga\Descargas Maestros Endpoints\listado_terceros (2).xlsx'    # Archivo con nueva clientes

# Leer el archivo 2 con los datos fuente
df2 = pd.read_excel(archivo_2)
columnas_a_insertar = ['NIT/CC', 'NOMBRE', '1ER_APELLIDO', '2DO_APELLIDO', 'ABREVIATURA', 'REGIMEN']
df2 = df2[columnas_a_insertar].fillna('')  # Reemplazar NaN por cadenas vacías

#Leer el archivo 3 con los datos fuente
df3 = pd.read_excel(archivo_3)
columnas_a_insertar_clientes = ['NIT/CC','NOMBRE','REGIMEN']

# Abrir el archivo 1 sin perder su estructura
wb = load_workbook(archivo_1)
ws = wb.active  # o ws = wb["Sheet1"] si sabes el nombre exacto de la hoja

# Encontrar la primera fila vacía (verifica si la columna B está vacía)
fila_inicial = 2  # Asumiendo que fila 1 es encabezado
while ws.cell(row=fila_inicial, column=2).value is not None:
    fila_inicial += 1

# Escribir los datos de terceros conductores 2, 3, 4, 5 (B, C, D, E)
for _, fila_datos in df2.iterrows():
    ws.cell(row=fila_inicial, column=1).value = fila_datos['NIT/CC']
    ws.cell(row=fila_inicial, column=2).value = fila_datos['NOMBRE']
    ws.cell(row=fila_inicial, column=3).value = fila_datos['1ER_APELLIDO']
    ws.cell(row=fila_inicial, column=4).value = fila_datos['2DO_APELLIDO']
    ws.cell(row=fila_inicial, column=5).value = fila_datos['ABREVIATURA']
    ws.cell(row=fila_inicial, column=17).value = fila_datos['REGIMEN']
    fila_inicial += 1

#Escribir los datos de terceros clientes
for _, fila_datos in df3.iterrows():
    ws.cell(row=fila_inicial, column=1).value = fila_datos['NIT/CC']
    ws.cell(row=fila_inicial, column=2).value = fila_datos['NOMBRE']
    ws.cell(row=fila_inicial, column=17).value = fila_datos['REGIMEN']
    fila_inicial += 1

# Guardar cambios
wb.save(archivo_1)
print("¡Datos copiados al archivo 1 correctamente!")