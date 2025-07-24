import pandas as pd
from openpyxl import load_workbook

rutamaestro = r'C:\Users\Mateo Zapata\OneDrive\Escritorio\Automatizaciones\2. Pedidos y Órdenes de Carga\Maestro Endpoints.xlsx'
rutaterceros = r'C:\Users\Mateo Zapata\OneDrive\Escritorio\Automatizaciones\2. Pedidos y Órdenes de Carga\Descargas Maestros Endpoints\listado_terceros.xlsx'

# Leer los archivos
df_maestro = pd.read_excel(rutamaestro, sheet_name='Asignar')
df_terceros = pd.read_excel(rutaterceros)

# Limpiar nombres de columnas
df_maestro.columns = df_maestro.columns.str.strip()
df_terceros.columns = df_terceros.columns.str.strip()

# Convertir claves a texto para evitar errores de tipo
df_maestro['cod_propie'] = df_maestro['cod_propie'].astype(str)
df_maestro['cod_tenedo'] = df_maestro['cod_tenedo'].astype(str)
df_maestro['conductor'] = df_maestro['conductor'].astype(str)

df_terceros['NIT/CC'] = df_terceros['NIT/CC'].astype(str)

# Crear un diccionario tipo BUSCARV: {clave: valor}
diccionario_lookup_propietario = dict(zip(df_terceros['NIT/CC'], df_terceros['Nombre Completo 2']))
diccionario_lookup_tenedor = dict(zip(df_terceros['NIT/CC'], df_terceros['Nombre Completo 2']))
diccionario_lookup_conductor_mayus = dict(zip(df_terceros['NIT/CC'], df_terceros['Nombre completo']))
diccionario_lookup_conductor_minus = dict(zip(df_terceros['NIT/CC'], df_terceros['Nombre Completo 2']))
diccionario_lookup_regimen = dict(zip(df_terceros['NIT/CC'], df_terceros['COD_REGIMEN']))

# Abrir el archivo completo sin perder hojas
libro = load_workbook(rutamaestro)
hoja = libro['Asignar']

# Recorrer las filas y escribir en la columna C (columna 3) el nombre Propietario
for i, cod in enumerate(df_maestro['cod_propie'], start=2):  # empieza en la fila 2
    nombre = diccionario_lookup_propietario.get(cod, None)
    if nombre:
        hoja.cell(row=i, column=3).value = nombre  # columna C = 3

# Recorrer las filas y escribir en la columna C (columna 5) el nombre Propietario
for i, cod in enumerate(df_maestro['cod_tenedo'], start=2):  # empieza en la fila 2
    nombre = diccionario_lookup_tenedor.get(cod, None)
    if nombre:
        hoja.cell(row=i, column=5).value = nombre  # columna C = 5

# Recorrer las filas y escribir en la columna C (columna 8) el nombre Conductor en Mayúscula
for i, cod in enumerate(df_maestro['conductor'], start=2):  # empieza en la fila 2
    nombre = diccionario_lookup_conductor_mayus.get(cod, None)
    if nombre:
        hoja.cell(row=i, column=8).value = nombre  # columna C = 8

# Recorrer las filas y escribir en la columna C (columna 9) el nombre Conductor en Minúscula
for i, cod in enumerate(df_maestro['conductor'], start=2):  # empieza en la fila 2
    nombre = diccionario_lookup_conductor_minus.get(cod, None)
    if nombre:
        hoja.cell(row=i, column=9).value = nombre  # columna C = 9

# Recorrer las filas y escribir en la columna C (columna 29) el Regimen
for i, cod in enumerate(df_maestro['cod_propie'], start=2):  # empieza en la fila 2
    nombre = diccionario_lookup_regimen.get(cod, None)
    if nombre:
        hoja.cell(row=i, column=29).value = nombre  # columna C = 29

# Guardar sin perder las otras hojas
libro.save(rutamaestro)

print('Se actualizan los nombres de propietario, tenedor, conductor y código de regimén')

