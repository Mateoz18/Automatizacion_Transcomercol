import pandas as pd
from openpyxl import load_workbook

rutamaestro = r'C:\Users\Mateo Zapata\OneDrive\Escritorio\Automatizaciones\2. Pedidos y Órdenes de Carga\Maestro Endpoints.xlsx'
rutamarcas = r'C:\Users\Mateo Zapata\OneDrive\Escritorio\Automatizaciones\2. Pedidos y Órdenes de Carga\Descargas Maestros Endpoints\Lista_Marcas.xlsx'
rutalineas = r'C:\Users\Mateo Zapata\OneDrive\Escritorio\Automatizaciones\2. Pedidos y Órdenes de Carga\Descargas Maestros Endpoints\Lista_Lineas.xlsx'
rutacolores = r'C:\Users\Mateo Zapata\OneDrive\Escritorio\Automatizaciones\2. Pedidos y Órdenes de Carga\Descargas Maestros Endpoints\Lista_Colores.xlsx'
rutacarrocerias = r'C:\Users\Mateo Zapata\OneDrive\Escritorio\Automatizaciones\2. Pedidos y Órdenes de Carga\Descargas Maestros Endpoints\Lista_Carrocerias.xlsx'
rutaopeGPS = r'C:\Users\Mateo Zapata\OneDrive\Escritorio\Automatizaciones\2. Pedidos y Órdenes de Carga\Descargas Maestros Endpoints\Lista_Operadores_GPS.xlsx'

# Leer los archivos
df_maestro = pd.read_excel(rutamaestro, sheet_name='Asignar')
df_marcas = pd.read_excel(rutamarcas)
df_lineas = pd.read_excel(rutalineas)
df_colores = pd.read_excel(rutacolores)
df_carrocerias = pd.read_excel(rutacarrocerias)
df_opegps = pd.read_excel(rutaopeGPS)

# Limpiar nombres de columnas
df_maestro.columns = df_maestro.columns.str.strip()
df_marcas.columns = df_marcas.columns.str.strip()
df_lineas.columns = df_lineas.columns.str.strip()
df_colores.columns = df_colores.columns.str.strip()
df_carrocerias.columns = df_carrocerias.columns.str.strip()
df_opegps.columns = df_opegps.columns.str.strip()

# Convertir claves a texto para evitar errores de tipo
df_maestro['nom_marcax'] = df_maestro['nom_marcax'].astype(str).str.strip()
df_maestro['nom_lineax'] = df_maestro['nom_lineax'].astype(str).str.strip()
df_maestro['nom_colorx'] = df_maestro['nom_colorx'].astype(str).str.strip()
df_maestro['nom_carroc'] = df_maestro['nom_carroc'].astype(str).str.strip()
df_maestro['cod_opegps'] = df_maestro['cod_opegps'].astype(str).str.strip()

df_marcas['MARCA'] = df_marcas['MARCA'].astype(str)
df_lineas['LINEA'] = df_lineas['LINEA'].astype(str)
df_colores['COLOR'] = df_colores['COLOR'].astype(str)
df_carrocerias['Nombre'] = df_carrocerias['Nombre'].astype(str)
df_opegps['OPERADOR GPS'] = df_opegps['OPERADOR GPS'].astype(str)

df_maestro['num_trayle'] = df_maestro['num_trayle'].fillna(0).replace('', 0)

# Crear un diccionario tipo BUSCARV: {clave: valor}
diccionario_lookup_marcas = dict(zip(df_marcas['MARCA'], df_marcas['CODIGO MARCA']))
diccionario_lookup_lineas = dict(zip(df_lineas['LINEA'], df_lineas['CODIGO LINEA']))
diccionario_lookup_colores = dict(zip(df_colores['COLOR'], df_colores['CODIGO COLOR']))
diccionario_lookup_carrocerias = dict(zip(df_carrocerias['Nombre'], df_carrocerias['Codigo']))
diccionario_lookup_opegps = dict(zip(df_opegps['OPERADOR GPS'], df_opegps['NIT']))

# Abrir el archivo completo sin perder hojas
libro = load_workbook(rutamaestro)
hoja = libro['Asignar']

# Recorrer las filas y escribir en la columna C (columna 13) la marca
for i, cod in enumerate(df_maestro['nom_marcax'], start=2):  # empieza en la fila 2
    nombre = diccionario_lookup_marcas.get(cod, None)
    if nombre:
        hoja.cell(row=i, column=13).value = nombre  # columna C = 13

# Recorrer las filas y escribir en la columna C (columna 15) la linea
for i, cod in enumerate(df_maestro['nom_lineax'], start=2):  # empieza en la fila 2
    nombre = diccionario_lookup_lineas.get(cod, None)
    if nombre:
        hoja.cell(row=i, column=15).value = nombre  # columna C = 15

# Recorrer las filas y escribir en la columna C (columna 17) el color
for i, cod in enumerate(df_maestro['nom_colorx'], start=2):  # empieza en la fila 2
    nombre = diccionario_lookup_colores.get(cod, None)
    if nombre:
        hoja.cell(row=i, column=17).value = nombre  # columna C = 17

# Recorrer las filas y escribir en la columna C (columna 21) la carroceria
for i, cod in enumerate(df_maestro['nom_carroc'], start=2):  # empieza en la fila 2
    nombre = diccionario_lookup_carrocerias.get(cod, None)
    if nombre:
        hoja.cell(row=i, column=21).value = nombre  # columna C = 21

# Recorrer las filas y escribir en la columna C (columna 26) el operador GPS
for i, cod in enumerate(df_maestro['cod_opegps'], start=2):  # empieza en la fila 2
    nombre = diccionario_lookup_opegps.get(cod, None)
    if nombre:
        hoja.cell(row=i, column=26).value = nombre  # columna C = 26
    else:
        hoja.cell(row=i, column=25).value = 0

# Escribir la columna num_trayle (posición 24) en la hoja Asignar
for i, valor in enumerate(df_maestro['num_trayle'], start=2):  # empieza en la fila 2
    hoja.cell(row=i, column=24).value = valor

# Guardar sin perder las otras hojas
libro.save(rutamaestro)

print('Se actualizaron los atributos: Líneas, Marcas, Colores, Carrocerías, Operador GPS')