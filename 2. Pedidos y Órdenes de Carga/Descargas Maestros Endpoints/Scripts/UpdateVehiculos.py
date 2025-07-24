import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
import xlwings as xw
from openpyxl.styles import numbers

# Rutas de los archivos
archivo_1 = r"C:\Users\Mateo Zapata\OneDrive\Escritorio\Automatizaciones\2. Pedidos y Órdenes de Carga\Maestro Endpoints.xlsx"

hoy = datetime.now().strftime("%Y_%m_%d")
archivo_2 = fr"C:\Users\Mateo Zapata\OneDrive\Escritorio\Automatizaciones\2. Pedidos y Órdenes de Carga\Descargas Maestros Endpoints\vehiculos_{hoy}.xlsx"

# Leer archivo 2 (vehiculos)
df2 = pd.read_excel(archivo_2)

# Crear dataframe para hoja 'Asignar' del archivo 1 con columnas necesarias
columnas_destino = [
    'placa', 'cod_propie', 'nom_propie', 'cod_tenedo', 'nom_tenedo', 
    'conductor', 'nconduc', 'nom_conduc', 'lab_conduc', 'capaci', 
    'Fecha Vencimiento SOAT', 'nom_marcax', 'cod_marcax', 'nom_lineax', 'cod_lineax', 
    'nom_colorx', 'cod_colorx', 'ano_modelo', 'ano_repote', 'nom_carroc', 
    'cod_carroc', 'num_config', 'val_pesmax', 'num_trayle', 'cod_opegps', 
    'nit_opegps', 'val_pesove', 'val_capaci', 'cod_terreg'
]

# Construir el DataFrame de salida usando el mapeo especificado
df_asignar = pd.DataFrame({
    'placa': df2['Placa'],
    'cod_propie': df2['C.C. Propietario'],
    'nom_propie': "", # Vacío
    'cod_tenedo': df2['C.C. Tenedor'],
    'nom_tenedo': "", # Vacío
    'conductor': df2['C.C. Conductor'],
    'nconduc': df2['Nombre Conductor'],
    'nom_conduc': "",  # Vacío
    'lab_conduc': "",  # Vacío
    'capaci': df2['Capacidad'],
    'Fecha Vencimiento SOAT': df2['Fecha Vencimiento SOAT'],
    'nom_marcax': df2['Marca'],
    'cod_marcax': "",  # Vacío
    'nom_lineax': df2['Línea'],
    'cod_lineax': "",  # Vacío
    'nom_colorx': df2['Color'],
    'cod_colorx': "",  # Vacío
    'ano_modelo': df2['Modelo'],
    'ano_repote': df2['Repotenciado'],
    'nom_carroc': df2['Carrocería'],
    'cod_carroc': "",  # Vacío
    'num_config': df2['Configuración'],
    'val_pesmax': "",  # Lo agregamos luego como fórmula
    'num_trayle': df2['Remolque'],
    'cod_opegps': df2['Operador GPS'],
    'nit_opegps': "",  # Vacío
    'val_pesove': df2['Peso Vacío (Tn)'],
    'val_capaci': df2['Capacidad'],
    'cod_terreg': ""  # Vacío
})

# Escribir en la hoja 'Asignar' del archivo 1
with pd.ExcelWriter(archivo_1, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    df_asignar.to_excel(writer, sheet_name='Asignar', index=False)


#Eliminar registros con fecha de SOAT inferior a 2025/01/01
maestro = r"C:\Users\Mateo Zapata\OneDrive\Escritorio\Automatizaciones\2. Pedidos y Órdenes de Carga\Maestro Endpoints.xlsx"

# Leer solo la hoja 'Asignar' con pandas
df = pd.read_excel(maestro, sheet_name="Asignar", engine="openpyxl")
'''
# Convertir columna de fecha (K) al tipo datetime si no lo está ya
df['fecha_columna_k'] = pd.to_datetime(df.iloc[:, 10], errors='coerce')  # K = columna 11 = index 10

# Filtrar por fecha
fecha_corte = pd.Timestamp("2025-01-01")
df_filtrado = df[df['fecha_columna_k'] >= fecha_corte]

# Eliminar la columna auxiliar si quieres
df_filtrado = df_filtrado.drop(columns=['fecha_columna_k'])

# Cargar el archivo sin perder otras hojas
wb = load_workbook(maestro)
with pd.ExcelWriter(maestro, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    df_filtrado.to_excel(writer, sheet_name='Asignar', index=False)
'''
#Agregar fórmula BUSCARV a la columna 'val_pesmax'
wb = load_workbook(maestro)
ws = wb['Asignar']

for row in range(2, ws.max_row + 1):
    ws[f"W{row}"] = f'=VLOOKUP(V{row},PBV!A:D,4,0)'  # Columna W = val_pesmax, Columna V = num_config

# Aplicar formato de fecha corta a la columna 'Fecha Vencimiento SOAT' (columna K)
for row in range(2, ws.max_row + 1):
    cell = ws[f"K{row}"]
    if isinstance(cell.value, datetime):
        cell.number_format = 'd/m/yyyy'
        
wb.save(maestro)



print("Información de vehículos actualizada exitosamente")

