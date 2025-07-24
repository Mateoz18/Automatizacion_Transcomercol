import pandas as pd
import shutil
import os
from datetime import datetime
import glob
from openpyxl import load_workbook

# 1. Copiar archivo Maestro Endpoints.xlsx a carpeta Histórico
ruta_base = r"C:\Users\Mateo Zapata\OneDrive\Escritorio\Automatizaciones\2. Pedidos y Órdenes de Carga"
ruta_destino = r"C:\Users\Mateo Zapata\OneDrive\Escritorio\Automatizaciones\2. Pedidos y Órdenes de Carga\Descargas Maestros Endpoints"
archivo_maestro = os.path.join(ruta_base, "Maestro Endpoints.xlsx")
destino_hist = os.path.join(ruta_destino, "Historicos", f"Maestro Endpoints copia {datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")

shutil.copy(archivo_maestro, destino_hist)
print("Archivo copiado a Histórico.")

# 2. Limpiar hoja 'Asignar' desde fila 2 usando openpyxl
wb = load_workbook(archivo_maestro)
ws = wb['Asignar']
ws.delete_rows(2, ws.max_row)
wb.save(archivo_maestro)
print("Hoja 'Asignar' limpiada correctamente (con openpyxl).")